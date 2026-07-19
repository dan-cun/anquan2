from __future__ import annotations

import hashlib
import re
from pathlib import Path

from openpyxl import load_workbook

from app.schemas.agents import AgentRole
from app.schemas.prompts import PromptMessageRole
from prompts.catalog import PromptDefinition

_ACTION_RE = re.compile(r"{{[\s\S]*?}}")
_VARIABLE_RE = re.compile(r"\.[A-Za-z_][A-Za-z0-9_]*")
_BLOCK_OPENERS = {"if", "range", "with", "define", "block"}


class PromptWorkbookError(ValueError):
    pass


def _variables(content: str) -> set[str]:
    return {
        variable
        for action in _ACTION_RE.findall(content)
        for variable in _VARIABLE_RE.findall(action)
    }


def _balanced_blocks(content: str) -> bool:
    depth = 0
    for action in _ACTION_RE.findall(content):
        body = action[2:-2].strip().strip("-").strip()
        directive = body.split(maxsplit=1)[0] if body else ""
        if directive in _BLOCK_OPENERS:
            depth += 1
        elif directive == "end":
            depth -= 1
            if depth < 0:
                return False
    return depth == 0


class PromptWorkbookImporter:
    sheet_name = "Prompt修改表"
    required_headers = {
        "Prompt键",
        "分类",
        "消息角色",
        "Agent/模块",
        "源文件",
        "原始Prompt（参考）",
        "修改后Prompt（编辑）",
        "修改状态",
    }

    def load(self, workbook_path: Path) -> list[PromptDefinition]:
        path = workbook_path.expanduser().resolve()
        if not path.is_file():
            raise PromptWorkbookError(f"Prompt workbook not found: {path}")
        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            if self.sheet_name not in workbook.sheetnames:
                raise PromptWorkbookError(f"Missing worksheet: {self.sheet_name}")
            worksheet = workbook[self.sheet_name]
            rows = list(worksheet.iter_rows(values_only=True))
        finally:
            workbook.close()

        header_index = next(
            (
                index
                for index, row in enumerate(rows)
                if "Prompt键" in {str(value) for value in row if value is not None}
            ),
            None,
        )
        if header_index is None:
            raise PromptWorkbookError("Prompt header row was not found")
        headers = [str(value or "").strip() for value in rows[header_index]]
        missing_headers = self.required_headers.difference(headers)
        if missing_headers:
            raise PromptWorkbookError(
                f"Prompt workbook is missing columns: {', '.join(sorted(missing_headers))}"
            )
        columns = {header: headers.index(header) for header in headers if header}

        definitions: list[PromptDefinition] = []
        keys: set[str] = set()
        for row in rows[header_index + 1 :]:
            key = str(row[columns["Prompt键"]] or "").strip()
            if not key:
                continue
            if key in keys:
                raise PromptWorkbookError(f"Duplicate Prompt key: {key}")
            keys.add(key)
            original = str(row[columns["原始Prompt（参考）"]] or "")
            modified = str(row[columns["修改后Prompt（编辑）"]] or "")
            if not modified.strip():
                raise PromptWorkbookError(f"Modified Prompt is empty: {key}")
            missing_variables = sorted(_variables(original).difference(_variables(modified)))
            if missing_variables:
                raise PromptWorkbookError(
                    f"Prompt {key} removed variables: {', '.join(missing_variables)}"
                )
            if not _balanced_blocks(modified):
                raise PromptWorkbookError(f"Prompt has unbalanced Go Template blocks: {key}")
            role_value = str(row[columns["消息角色"]] or "template").strip().lower()
            try:
                message_role = PromptMessageRole(role_value)
            except ValueError as error:
                raise PromptWorkbookError(
                    f"Prompt {key} has invalid message role: {role_value}"
                ) from error
            try:
                agent_role = AgentRole(key)
            except ValueError:
                agent_role = None
            variables_value = (
                str(row[columns.get("模板变量", -1)] or "")
                if "模板变量" in columns
                else ""
            )
            definitions.append(
                PromptDefinition(
                    key=key,
                    content=modified,
                    category=str(row[columns["分类"]] or "native"),
                    message_role=message_role,
                    name=str(row[columns["Agent/模块"]] or key),
                    source_path=str(row[columns["源文件"]] or "") or None,
                    variables=[item for item in variables_value.splitlines() if item],
                    agent_role=agent_role,
                    metadata={
                        "workbook": path.name,
                        "workbook_status": str(row[columns["修改状态"]] or ""),
                        "workbook_notes": str(
                            row[columns.get("修改说明", -1)] or ""
                            if "修改说明" in columns
                            else ""
                        ),
                        "locale": "zh-CN",
                    },
                    checksum=hashlib.sha256(modified.encode("utf-8")).hexdigest(),
                )
            )
        if len(definitions) != 41:
            raise PromptWorkbookError(f"Expected 41 Prompts, found {len(definitions)}")
        return definitions
